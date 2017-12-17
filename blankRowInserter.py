#このプログラムは任意のExcelファイルに改行を好きなだけ挿入するプログラムです
#blankRowInserter.py　改行を挿入する行　改行の量　対象ファイル
import sys
import openpyxl

#改行を挿入する行，改行量，対象ファイルをコマンドラインから取得するようにする
blank_insert_start=int(sys.argv[1])
blank_insert_volum=int(sys.argv[2])
file_name=str(sys.argv[3])

#対象ファイルのオブジェクトとシート変数を生成
wb=openpyxl.load_workbook(file_name)
sheet=wb.active

#新しいファイルを生成
wb_2=openpyxl.Workbook()
sheet_2=wb_2.active

#blank_insert_start行目まではファイルのコピーと一緒
for row_num in range(1,blank_insert_start):
    for column_num in range(1,sheet.max_column+1):
        sheet_2.cell(row=row_num,column=column_num).value=sheet.cell(row=row_num,column=column_num).value

#blank_insert_start行目からblank_insert_volumだけ空白行を挿入
for row_num in range(blank_insert_start,blank_insert_start+blank_insert_volum):
    for column_num in range(1,sheet.max_column+1):
        sheet_2.cell(row=row_num,column=column_num).value=""

#blank_insert_start＋lank_insert_volumから最後までは元の入力値をコピペ
for row_num in range(blank_insert_start+blank_insert_volum,sheet.max_row+blank_insert_volum+1):
    for column_num in range(1,sheet.max_column+1):
        sheet_2.cell(row=row_num,column=column_num).value=sheet.cell(row=row_num-blank_insert_volum,column=column_num).value

#blankRowInserter.xlsxとしてファイルを保存
wb_2.save('blankRowInserter.xlsx')
